'''
Created on 2019年7月21日

@author: 55057
'''



import datetime
import openpyxl
import logging
from book_querier.book_querier_engine import ISBNSearchEngine
from ORM_model.decBaseClass import Book
from dispatcher.rawdata_orm import InsertBatchBooks



#从Execl获取数据
def execlDataToList(execl_filepath):
        #打开工作簿
    workbook = openpyxl.load_workbook(execl_filepath)
    sheet = workbook.active
    logging.info('workbook row is:'+str(sheet.max_row))
    #定义ISBN列表
    list_isbn = []
    for row_id in range(1,sheet.max_row+1):
       #增加ISBN
       list_isbn.append(sheet["A"+str(row_id)].value)
        #输出信息
    logging.info('ISBN length is'+str(len(list_isbn)))
    print('ISBN length is:'+str(len(list_isbn)))
    return list_isbn

#通过ISBN查询图书详细信息
def ISBNListToBookInformation(list_isbn):
    #图书详细内容
    isbn_list_length = len(list_isbn)
    bookdetaillist = []
    isbnEngine = ISBNSearchEngine()
    for index in range(0,isbn_list_length):
        #查找图书详情
        bookdetail = isbnEngine.get_information_with_ISBN(list_isbn[index])
        if bookdetail == -1:
            #无此图书信息(占位补充)
            bookdetail = ['none','none','none','none','none']
        #构建图书详细信息
        bookdetaillist.append(bookdetail)
    logging.info('book detail list build succeed.')
    return bookdetaillist
    
#构建输入基础数据集
def BuildDatabaseRawDataORM(execl_filepath):
        #原始数据列表
        ORM_data_list = []
        #isbn清单
        list_isbn = execlDataToList(execl_filepath)
        #图书信息清单
        book_details_list = ISBNListToBookInformation(list_isbn)
        logging.info("build base information succeed.")
        #
        if len(list_isbn) != len(book_details_list):
            logging.error('列表数据长度不符')
            raise Exception('长度不符')
        for index in range(0,len(list_isbn)):
            #构建基础元组数据
            tempbook = Book(ISBN=list_isbn[index],book_name=book_details_list[index][0],book_author=book_details_list[index][1],
                    book_publisher=book_details_list[index][2],book_number=book_details_list[index][3],
                    book_price=book_details_list[index][4],book_notes=None,
                    book_storoge_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            
            #构建整体数据集
            ORM_data_list.append(tempbook)
    
        return ORM_data_list




#test function
if __name__ == '__main__':
    #构建ORM原型数据
    books = BuildDatabaseRawDataORM('D:/raw_data.xlsx');
    print(len(books))
    #图书批量插入
    InsertBatchBooks(books)
    



        
    
