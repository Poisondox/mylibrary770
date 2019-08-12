'''
Created on 2019�?8�?7�?

@author: 55057
'''
from sqlalchemy.engine import create_engine
from sqlalchemy.orm.session import sessionmaker
from sqlalchemy.orm.scoping import scoped_session
from ORM_model.decBaseClass import initDatabase, destoryDatabase, Book
from contextlib import contextmanager
import time
import datetime
import openpyxl
from core.book_querier_engine import ISBNSearchEngine
from core.get_config import getConfigFromFile
from core.loggers import logger

# 初始化数据库连接
sql_engine = create_engine(getConfigFromFile('db_cfg.ini'), echo=True)
# 初始化数据库实例session
sessionType = scoped_session(sessionmaker(bind=sql_engine))
# �?毁数据库
destoryDatabase(sql_engine)
# 初始化数据库
initDatabase(sql_engine)


# 获取session实例操作数据�?
def getSession():
    return sessionType()


# 数据通过ORM手段写入数据库，方法实现于此文件
@contextmanager
def session_flow():
    session = getSession()
    try:
        yield session
        session.commit()
    except:
        session.rollback()
        raise 
    finally:
        session.close()

        
# 从Execl获取数据
def execlDataToList(execl_filepath):
        # 打开工作�?
    workbook = openpyxl.load_workbook(execl_filepath)
    sheet = workbook.active
    logger.info('workbook row is:' + str(sheet.max_row))
    # 定义ISBN列表
    list_isbn = []
    for row_id in range(1, sheet.max_row + 1):
        # 增加ISBN
        list_isbn.append(sheet["A" + str(row_id)].value)
        # 输出信息
    logger.info('ISBN length is' + str(len(list_isbn)))
    return list_isbn


# 通过ISBN查询图书详细信息
def ISBNListToBookInformation(list_isbn):
    # 图书详细内容
    isbn_list_length = len(list_isbn)
    bookdetaillist = []
    # 初始化检索引�?
    isbnEngine = ISBNSearchEngine()
    for index in range(0, isbn_list_length):
        # 查找图书详情
        bookdetail = isbnEngine.get_information_with_ISBN(list_isbn[index])
        if bookdetail == -1:
            # 无此图书信息(占位补充)
            bookdetail = ['none', 'none', 'none', 'none', 'none']
        # 构建图书详细信息
        bookdetaillist.append(bookdetail)
    logger.info('book detail list build succeed.')
    return bookdetaillist

    
# 构建输入基础数据�?
def BuildDatabaseRawDataORM(execl_filepath):
        # 原始数据列表
        ORM_data_list = []
        # ISBN清单
        list_isbn = execlDataToList(execl_filepath)
        # 图书信息清单
        book_details_list = ISBNListToBookInformation(list_isbn)
        logger.info("build base information succeed.")
        #
        if len(list_isbn) != len(book_details_list):
            logger.error('列表数据长度不符')
            raise Exception('长度不符')
        for index in range(0, len(list_isbn)):
            # 输出日志信息
            logger.info('at present: ' + str(index + 1) + 'sum is :' + str(len(list_isbn)))
            # 构建基础元组数据
            tempbook = Book(ISBN=list_isbn[index], book_name=book_details_list[index][0], book_author=book_details_list[index][1],
                    book_publisher=book_details_list[index][2], book_number=book_details_list[index][3],
                    book_price=book_details_list[index][4], book_notes=None,
                    book_storoge_time=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            # 构建整体数据�?
            ORM_data_list.append(tempbook)
        return ORM_data_list

   
# ORM功能        
# 单本图书插入
def InsertBookData(isbn, ibook_name, ibook_author, ibook_publisher, ibook_number, ibook_price, ibook_notes, ibook_storoge_time):
    with session_flow() as temp_session:
        book = Book(ISBN=isbn, book_name=ibook_name, book_author=ibook_author,
                    book_publisher=ibook_publisher, book_number=ibook_number,
                    book_price=ibook_price, book_notes=ibook_notes,
                    book_storoge_time=ibook_storoge_time)
        # 录入图书数据
        temp_session.add(book)


def InsertBook(ibook):
    with session_flow() as temp_session:
        temp_session.add(ibook)


# 图书批量插入
def InsertBatchBooks(booklist):
    with session_flow() as temp_session:
        # 全部写入session
        temp_session.add_all(booklist)
        
        
if __name__ == '__main__':
    InsertBookData('1987564', '韭菜的自我修养', '阿波罗', '豌豆苗出版社', '自家出版社', 48.56, None, time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))        

